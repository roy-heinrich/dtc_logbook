import json
import os
import re
import sys
from copy import copy

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LOCAL_PYTHON_PACKAGES = os.path.join(PROJECT_ROOT, 'python_packages')

if os.path.isdir(LOCAL_PYTHON_PACKAGES) and LOCAL_PYTHON_PACKAGES not in sys.path:
    sys.path.insert(0, LOCAL_PYTHON_PACKAGES)

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font


DEFAULT_FONT = Font(name="Calibri", size=11)


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def format_sex(sex_value):
    return normalize_text(sex_value)


def sanitize_name(name_value):
    name = normalize_text(name_value)
    return re.sub(r"\s+(n/a|n/|na|n|-|none|null)\s*$", "", name, flags=re.IGNORECASE).strip()


def format_contact(email, number):
    email_text = normalize_text(email) or "-"
    number_text = normalize_text(number) or "-"
    return f"Email: {email_text}\nNumber: {number_text}"


def apply_header_image(sheet):
    header_candidates = [
        os.path.join(PROJECT_ROOT, '-header.png'),
        os.path.join(PROJECT_ROOT, 'public', 'images', 'header-banner.png'),
    ]

    header_path = next((path for path in header_candidates if os.path.isfile(path)), None)
    if not header_path:
        return

    if hasattr(sheet, '_images'):
        sheet._images = []

    sheet.row_dimensions[1].height = 95

    image = Image(header_path)
    image.width = 795
    image.height = 125
    sheet.add_image(image, 'A1')


def copy_row_style(ws, src_row, dst_row, max_col=8):
    for col in range(1, max_col + 1):
        source_cell = ws.cell(row=src_row, column=col)
        target_cell = ws.cell(row=dst_row, column=col)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        target_cell.number_format = source_cell.number_format
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)

    if src_row in ws.row_dimensions:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def main():
    if len(sys.argv) < 4:
        print("Usage: python generate_dtc_attendance.py <template_path> <output_path> <payload_path>")
        sys.exit(1)

    template_path = sys.argv[1]
    output_path = sys.argv[2]
    payload_path = sys.argv[3]

    with open(payload_path, "r", encoding="utf-8") as handle:
        payload = json.load(handle)

    attendees = payload.get("attendees", [])
    date_label = normalize_text(payload.get("date_label"))
    venue_label = normalize_text(payload.get("venue_label"))
    services_header = normalize_text(payload.get("services_header"))

    workbook = load_workbook(template_path)
    sheet = workbook.active

    apply_header_image(sheet)
    sheet["A2"] = f"Venue: {venue_label or 'DTC - DICT Aklan Provincial Field Office (Co - Working Space)'}"
    sheet["A3"] = f"Date: {date_label}"
    sheet["A4"] = "ATTENDANCE SHEET"

    sheet["A5"] = "Name"
    sheet["C5"] = "Sex"
    sheet["D5"] = "Age"
    sheet["E5"] = "Services Availed\n" + (services_header or "CWS, Conference Room, Training Room")
    sheet["F5"] = "Contact Information"
    sheet["G5"] = "Sector"
    sheet["H5"] = "Terms"

    for header_cell in ("A5", "C5", "D5", "E5", "F5", "G5", "H5"):
        sheet[header_cell].font = DEFAULT_FONT

    data_start_row = 6
    footer_start_row = 19
    template_row_for_style = 6
    base_capacity = footer_start_row - data_start_row

    extra_rows = max(0, len(attendees) - base_capacity)
    if extra_rows > 0:
        sheet.insert_rows(footer_start_row, amount=extra_rows)
        for row_index in range(footer_start_row, footer_start_row + extra_rows):
            copy_row_style(sheet, template_row_for_style, row_index)

    total_rows = max(base_capacity, len(attendees))

    for index in range(total_rows):
        row = data_start_row + index
        attendee = attendees[index] if index < len(attendees) else {}

        for column in ("A", "B", "C", "D", "E", "F", "G", "H"):
            sheet[f"{column}{row}"] = ""

        name = sanitize_name(attendee.get("name"))
        age_value = attendee.get("age")
        age = "" if age_value in (None, "") else str(age_value)
        service = normalize_text(attendee.get("service"))
        sex_value = format_sex(attendee.get("sex"))
        contact_value = format_contact(attendee.get("email"), attendee.get("number"))
        sector_value = normalize_text(attendee.get("sector"))
        terms_user = normalize_text(attendee.get("terms_user"))

        if not name:
            sex_value = ""
            age = ""
            service = ""
            contact_value = ""
            sector_value = ""
            terms_user = ""

        sheet[f"A{row}"] = name
        sheet[f"C{row}"] = sex_value
        sheet[f"D{row}"] = age
        sheet[f"D{row}"].number_format = "0"
        sheet[f"E{row}"] = service
        sheet[f"F{row}"] = contact_value
        sheet[f"G{row}"] = sector_value
        sheet[f"H{row}"] = terms_user

        for column in ("A", "B", "C", "D", "E", "F", "G", "H"):
            sheet[f"{column}{row}"].font = DEFAULT_FONT

    workbook.save(output_path)


if __name__ == "__main__":
    main()
